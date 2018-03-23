#coding=utf-8

from django.shortcuts import render,HttpResponseRedirect,redirect,HttpResponse
from .models import  *
from .forms import *
from django.http import JsonResponse
from  django.core import serializers
from django.db.models import Q
import  openpyxl
from openpyxl import styles
import  os
import json
from django.core.paginator import *
from datetime import *

def login_form(req):

    return render(req,'login.html',locals())

# Create your views here.
#展示店铺信息并增加修改店铺信息
def shop(request):
     shop_list=ShopInfo.objects.all()
     man_list=Managers.objects.all()
     area_list=Area.objects.all()
     return render(request,'shop.html',locals())
#添加店铺信息


def addinfo(req):
    sava_message = {"sava_message": "保存成功"}
    action=req.GET['action']
   #判断添加信息类型：manager--添加人员信息 area--添加区域信息 shopinfo--添加店铺信息
   #添加人员信息
    if action=='manager':
        Managers.objects.create(
            name=req.GET['manager_name'],
            personal_cellphone=req.GET['manager_phone'],
            company_cellphone=req.GET['manager_c_phone'],
            qq=req.GET['manager_qq'],
            title=req.GET['manager_title'],
            email=req.GET['manger_email'],
            address=req.GET['manager_address']
        )
     #添加区域信息
    elif action=='area':
        manager_name=Managers.objects.get(id=req.GET['area_manager'])
        Area.objects.create(
            manager=manager_name,
            name=req.GET["area_name"]
        )
     #添加店铺信息
    elif action=="shopinfo" :
        sys_name = req.GET['sys_name']
        s_name = req.GET['s_name']
        sale_manager = Managers.objects.get(id=req.GET['sale_manager'])
        aree = Area.objects.get(id=req.GET['area'])
        shoptype = req.GET['shoptype']
        malltype = req.GET['malltype']
        opendate = req.GET['openingdate']
        shopadd = req.GET['shopadd']
        conbengin = req.GET['contractBegindate']
        conEnd = req.GET['contractEndDate']
        shopstate = req.GET['shopstate']

        ShopInfo.objects.update_or_create(sysName=sys_name,
                                sName=s_name,
                                managerId=sale_manager,
                                areaId=aree,
                                shopType=shoptype,
                                mallType=malltype,
                                openingDate=opendate,
                                shopAddress=shopadd,
                                contractBeginDate=conbengin,
                                contractEndDate=conEnd,
                                state=shopstate)

    return JsonResponse(sava_message)

def addcheckdata(req):
    sava_message = {"sava_message": "提交成功"}
    shopname=ShopInfo.objects.get(Id=req.GET["shop_name"])
    data_date=req.GET["data_date"]
    shop_amount_report = float(req.GET["money"])
    #根据data_date和shopname查询是否存在记录，如存在则更新，否则新建
    if datadiff.objects.filter(date=data_date,id_shop=shopname).count():
        #获取原有记录中的sys_amount
        sys_amount_exists=datadiff.objects.get(date=data_date, id_shop=shopname).sys_amount
        #更新记录中的shop_amount和amount
        datadiff.objects.filter(date=data_date, id_shop=shopname).update(shop_amount=shop_amount_report,\
                                                                         amount=sys_amount_exists-shop_amount_report)
    else:

        create_list = datadiff(date=data_date, shop_amount=shop_amount_report, id_shop=shopname)
        create_list.save()
    return JsonResponse(sava_message)

def addsyscheckdata(req):
    sava_message = {"sava_message": "提交成功"}
    shopname=ShopInfo.objects.get(Id=req.GET["shop_name"])
    data_date=req.GET["data_date"]
    sys_amount_report = float(req.GET["money"])
    #根据data_date和shopname查询是否存在记录，如存在则更新，否则新建
    if datadiff.objects.filter(date=data_date,id_shop=shopname).count():
        #获取原有记录中的sys_amount
        shop_amount_exists=datadiff.objects.get(date=data_date, id_shop=shopname).shop_amount
        #更新记录中的shop_amount和amount
        datadiff.objects.filter(date=data_date, id_shop=shopname).update(sys_amount=sys_amount_report,\
                                                                         amount=sys_amount_report-shop_amount_exists)
    else:

        create_list = datadiff(date=data_date, shop_amount=sys_amount_report, id_shop=shopname)
        create_list.save()
    return JsonResponse(sava_message)

import calendar
import time
one_page_data=7
# @high_paginator(7)
def checkall(req):
    try:
        curPage = int(req.GET.get('curPage', 1))
        allPage = int(req.GET.get('allPage', 1))
        pageType = str(req.GET.get('pageType', ''))
    except:
        curPage = 1
        allPage = 1
    if pageType == 'p_down':
        curPage += 1
    elif pageType == 'p_up':
        curPage -= 1
    '''
    得到当月第一天的日期和最后一天的日期
    查询当月有差异的数据
    '''
    #获取当前时间
    day_now=time.localtime()
    #当月第一天日期
    day_begin='%d-%02d-01'%(day_now.tm_year,day_now.tm_mon)
    wday,monthRange=calendar.monthrange(day_now.tm_year,day_now.tm_mon)
    #当月最后一天日期
    day_end='%d-%02d-%02d'%(day_now.tm_year, day_now.tm_mon, monthRange)
    shop_list = ShopInfo.objects.values('Id', 'sysName','sName', 'shopType', "managerId__name", "managerId").filter(
        shopType__in=["D", "C"],state__in=['S']).order_by("sysName")

    man_list = Managers.objects.values('id', 'name').filter(shopinfo__shopType__in=["D", "C"]).distinct()

    startDiff = (curPage - 1) * one_page_data
    endDiff = startDiff + one_page_data

    diff = datadiff.objects.exclude(amount=0)\
                           .exclude(id_shop__sName__in=['元隆利嘉生活馆','满洲里友谊商厦'])\
                           .filter(id_shop__shopType__in=["D", "C"],
                                   date__range=(day_begin,day_end))\
                           .order_by("id_shop", "-date")[startDiff:endDiff]

    if curPage ==1 and allPage==1:
        diffcounts= datadiff.objects.exclude(amount=0).exclude(id_shop__sName__in=['元隆利嘉生活馆', '满洲里友谊商厦']) \
                            .filter(id_shop__shopType__in=["D", "C"],date__range=(day_begin, day_end)) \
                            .order_by("id_shop", "-date").count()
        allPage=diffcounts/one_page_data
        remainDiff=diffcounts % one_page_data
        if remainDiff>0:
            allPage+=1
    page_list=range(1,allPage+1)

    return render(req,'checkall.html',locals())
#获取所有有差异的数据
def checkdata(request):

    try:
        curPage=int(request.GET.get('curPage',1))
        allPage=int(request.GET.get('allPage',1))
        pageType=str(request.GET.get('pageType',''))
    except ValueError:
        curPage=1
        allPage=1
    if pageType=='p_down':
        curPage+=1
    elif pageType=='p_up':
        curPage-=1
    startDiff=(curPage-1)*one_page_data
    endDiff=startDiff+one_page_data
    shop_list = ShopInfo.objects.values('Id', 'sName', 'sysName','shopType', "managerId__name", "managerId").filter(
        shopType__in=["D", "C"],state__in=['S']).order_by("-sName")
    man_list=Managers.objects.values('id','name').filter(shopinfo__shopType__in=["D","C"]).distinct()
    sel_shop_name=request.GET['select_name']
    sel_man_name=request.GET['man_name']
    sel_none = request.GET['sel_none']
    sel_date=request.GET['sel_date_start']
    sel_date_end=request.GET['sel_date_end']

    #构造查询条件字典
        #查询所有有差异的记录
    if sel_shop_name == '' and sel_man_name == '' and sel_date == '' and sel_none == '' and sel_date_end == '':
        filter_dic={'id_shop__shopType__in':["D","C"]}
        #按门店查询
    elif sel_shop_name and sel_man_name=='' and sel_date=='' and sel_none=='' and sel_date_end=='':
        filter_dic={'id_shop__shopType__in':["D","C"],'id_shop__Id':sel_shop_name}
        #按销售经理查询
    elif sel_man_name and sel_shop_name=='' and sel_date=='' and sel_none=='' and sel_date_end=='':
        filter_dic = {'id_shop__shopType__in': ["D", "C"], 'id_shop__managerId': sel_man_name}
        # 按时间查询
    elif sel_date and sel_shop_name == '' and sel_man_name == '' and sel_none == '' and sel_date_end:
        filter_dic = {'id_shop__shopType__in':["D", "C"],'date__range':(sel_date, sel_date_end)}
         # 按店铺和时间查询
    elif sel_date and sel_date_end and sel_shop_name and sel_man_name == '' and sel_none == '':
        filter_dic = {'id_shop__shopType__in':["D", "C"],'date__range':(sel_date, sel_date_end),'id_shop__Id':sel_shop_name}
        # 按销售经理和时间查询
    elif sel_man_name and sel_date and sel_date_end and sel_shop_name == '' and sel_none == '':
        filter_dic = {'id_shop__shopType__in':['D', 'C'],'date__range':(sel_date, sel_date_end),'id_shop__managerId':sel_man_name}
        # 查询所有未核查的结果
    elif sel_none == '1' and sel_shop_name == '' and sel_man_name == '' and sel_date == '' and sel_date_end == '':
        filter_dic = {'id_shop__shopType__in':['D', 'C'],'remark':u'未核查'}
        # 查询某段时间内的未核查结果
    elif sel_none == '1' and sel_date and sel_date_end and sel_man_name == '' and sel_shop_name == '':
        filter_dic = {'id_shop__shopType__in':['D', 'C'],'remark':u'未核查','date__range':(sel_date, sel_date_end)}
         # 查询销售经理未核查的结果
    elif sel_none == '1' and sel_date == '' and sel_man_name and sel_shop_name == '':
        filter_dic = {'id_shop__shopType__in':['D', 'C'],'remark':u'未核查','id_shop__managerId':sel_man_name}
        # 查询单店未核查的结果
    elif sel_none == '1' and sel_date == '' and sel_man_name == '' and sel_shop_name and sel_date_end == '':
        filter_dic = {'id_shop__shopType__in':['D', 'C'],'remark':u'未核查','id_shop__Id':sel_shop_name}
        # 查询单店给定时间段内的结果
    elif sel_date_end and sel_date and sel_shop_name and sel_man_name == '' and sel_none == '':
        filter_dic={'id_shop__shopType__in':['D', 'C'],'id_shop__Id' :sel_shop_name,'date__range' : (sel_date, sel_date_end)}
    # 构造查询条件字典

    #依据查询字典查询记录
    diff=datadiff.objects.exclude(amount=0)\
                         .exclude(id_shop__sName__in=['元隆利嘉生活馆','满洲里友谊商厦'])\
                         .filter(**filter_dic)\
                         .order_by("id_shop","-date")[startDiff:endDiff]
    #计算总分页数
    if curPage == 1 and allPage == 1:
        diffcounts = datadiff.objects.exclude(amount=0) \
            .exclude(id_shop__sName__in=['元隆利嘉生活馆', '满洲里友谊商厦']) \
            .filter(**filter_dic) \
            .order_by("id_shop", "-date").count()
        allPage = diffcounts / one_page_data
        remainDiff = diffcounts % one_page_data
        if remainDiff > 0:
            allPage += 1
    page_list=range(1,allPage+1)
    return render(request,'checkdatanew.html',locals())
#差异Excel导出
def diff_export_excel(request):
    sel_date_start=request.GET['date_start']
    sel_date_end=request.GET["date_end"]
    diff = datadiff.objects\
        .exclude(amount=0)\
        .exclude(id_shop__sName__in=['元隆利嘉生活馆','满洲里友谊商厦'])\
        .filter(id_shop__shopType__in=["D", "C"], date__range=(sel_date_start,sel_date_end))\
        .order_by("id_shop", "date")
    # 创建一个excel表格对象
    F_styleB=styles.Font(color=styles.colors.BLACK,bold=True)
    F_styleR=styles.Font(color=styles.colors.RED,bold=True,italic=True)
    F_styleCenter=styles.Alignment(horizontal='center',vertical='center',indent=1,readingOrder=2,text_rotation=0)

    diff_excel_book = openpyxl.Workbook()
    diff_excel_book.remove_sheet(diff_excel_book.get_sheet_by_name(u'Sheet'))
    book_sheet = diff_excel_book.create_sheet(u"销售差异")
    book_sheet.append([u"店铺",u'销售经理', u"日期", u"系统金额", u"上报金额", u"差异金额",u'正确金额', u"差异原因", u"备注"])
    for i in range(1,8):
        book_sheet.cell(row=1,column=i).font=F_styleB
    row=2
    for diff_y in diff:
            book_sheet.cell(row=row, column=1).value = diff_y.id_shop.sysName
            book_sheet.cell(row=row, column=1).alignment=F_styleCenter
            book_sheet.cell(row=row, column=2).value = diff_y.id_shop.managerId.name
            book_sheet.cell(row=row, column=2).alignment = F_styleCenter
            book_sheet.cell(row=row, column=3).value = diff_y.date
            book_sheet.cell(row=row, column=3).alignment = F_styleCenter
            book_sheet.cell(row=row, column=4).value = diff_y.sys_amount
            book_sheet.cell(row=row, column=4).alignment = F_styleCenter
            book_sheet.cell(row=row, column=5).value = diff_y.shop_amount
            book_sheet.cell(row=row, column=5).alignment = F_styleCenter
            book_sheet.cell(row=row, column=6).value = diff_y.amount
            book_sheet.cell(row=row, column=6).alignment = F_styleCenter
            book_sheet.cell(row=row, column=7).value = diff_y.true_amount
            if diff_y.true_amount!=diff_y.shop_amount:
                book_sheet.cell(row=row, column=7).font = F_styleR
            book_sheet.cell(row=row, column=7).alignment = F_styleCenter
            book_sheet.cell(row=row, column=8).value = diff_y.diff
            book_sheet.cell(row=row, column=8).alignment=F_styleCenter
            book_sheet.cell(row=row, column=9).value = diff_y.remark
            book_sheet.cell(row=row, column=9).alignment = F_styleCenter
            row+=1
    month_start = str(sel_date_start).replace('-','')
    month_end = str(sel_date_end).replace('-','')
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=销售差异%s-%%s.xlsx'%month_start %month_end
    diff_excel_book.save(response)
    return  response

#修改差异原因

def update_diff(req):
    sava_message = {"sava_message": "提交成功"}
    diff_new=req.POST["diff_new"]
    remark_new = req.POST["remark"]
    true_amount=req.POST["true_amount"]
    diff_id=req.POST["diffid"]
    datadiff.objects.filter(id=diff_id).update(diff=diff_new, remark=remark_new, true_amount=true_amount)
    # if true_amount=="sys":
    #         datadiff.objects.filter(id=diff_id).update(diff=diff_new,remark=remark_new,true_amount=1)
    # elif true_amount=="shop":
    #         datadiff.objects.filter(id=diff_id).update(diff=diff_new, remark=remark_new,true_amount=0)
    # elif true_amount=="show":
    #     datadiff.objects.filter(id=diff_id).update(diff=diff_new, remark=remark_new,true_amount=3)
    return  JsonResponse(sava_message)

#导入excel读写库 openpyxl(只可读写2007版及之后的office文档)
def changediff(req):
    if req.method=='POST':
        post_dict=req.POST
    id_pk=post_dict['pk']
    value_post=post_dict['value']
    name_post=post_dict['name']
    if 'remark' in name_post:
        datadiff.objects.filter(id=id_pk).update(remark=value_post)
    elif 'diff' in name_post:
        datadiff.objects.filter(id=id_pk).update(diff=value_post)
    elif 'true_amount' in name_post:
        datadiff.objects.filter(id=id_pk).update(true_amount=value_post)
    return  HttpResponse('YES')
def excelindb(request):
    import os
    #店铺列表
    shop_list = ShopInfo.objects.values('sysName','Id', 'sName', 'shopType', "managerId__name", "managerId").filter(
        shopType__in=["D", "C"]).order_by("-sName")
    #销售经理列表
    man_list = Managers.objects.values('id', 'name').filter(shopinfo__shopType__in=["D", "C"]).distinct()
    sava_message = {"sava_message": "上传成功"}
    excel_file = request.FILES['excelfile']
    print (excel_file)
    excel_date = request.POST["exceldate"]
    excel_data_type=request.POST['upinfo_type']
    #获取项目根目录
    filepath=os.path.abspath(os.path.dirname('__file__'))
    with open(filepath +'/static/upload/'+excel_data_type+ excel_date + '.xlsx', 'wb+') as destination:
        for chunk in excel_file.chunks():
            destination.write(chunk)

    destination.close()
    excel_data = openpyxl.load_workbook(filepath+'/static/upload/'+ excel_data_type+excel_date+".xlsx")
    sheetnames = excel_data.get_sheet_names()
    excel_sheets=excel_data.get_sheet_by_name(sheetnames[0])
    excel_rows=excel_sheets.max_row
    excel_columns=excel_sheets.max_column
    date_excel_start=(excel_sheets.cell(row=1,column=2).value).strftime("%Y-%m-%d")
    #获取当前人员表总记录数数
    if excel_data_type=="sysamout":
        for row_excel in range(2, excel_rows + 1):
            #循环读取每行数据然后写入数据库
            for column_excel in range(2,excel_columns+1):
                #excel表格从第二行开始读取，每行第一列为店铺名
                shop_id=ShopInfo.objects.get(sysName=(excel_sheets.cell(row=row_excel,column=1).value))
                date_excel = excel_sheets.cell(row=1, column=column_excel).value
                #每行第二列为系统金额
                sys_amount_excel=float(excel_sheets.cell(row=row_excel,column=column_excel).value)
                #根据时间（excel_date）和店铺名称（shop_id）来判断记录是否存在，如果存在更新原有记录，否则新建记录
                if datadiff.objects.filter(date=date_excel, id_shop=shop_id).count():
                    shop_amount_base = datadiff.objects.get(date=date_excel, id_shop=shop_id).shop_amount
                    datadiff.objects.filter(date=date_excel, id_shop=shop_id).update(sys_amount=sys_amount_excel,\
                                                                                     amount=sys_amount_excel-shop_amount_base,\
                                                                                     true_amount=shop_amount_base)
                else:
                    sys_diff=datadiff(sys_amount=sys_amount_excel,date=date_excel,id_shop=shop_id,true_amount=sys_amount_excel)
                    sys_diff.save()
                date_excel_str=date_excel.strftime("%Y-%m-%d")
        diff = datadiff.objects.exclude(amount=0) \
                    .exclude(id_shop__sName__in=['元隆利嘉生活馆', '满洲里友谊商厦']) \
                    .filter(id_shop__shopType__in=["D", "C"],
                            date__range=(date_excel_start, date_excel_str)) \
                    .order_by("id_shop", "date")

        os.remove(filepath+'/static/upload/'+excel_data_type+excel_date+".xlsx")
        return render(request,'checkdatanew.html',{"diff":diff,"shop_list":shop_list,"man_list":man_list})
    elif excel_data_type=="shopamout":
        for row_excel in range(2, excel_rows + 1):
                for column_excel in range(2,excel_columns+1):
                #excel表格从第二行开始读取，每行第一列为店铺名
                    shop_id=ShopInfo.objects.get(sName=(excel_sheets.cell(row=row_excel,column=1).value))
                    #每行第二列为店铺金额

                    shopamount_excel=float(excel_sheets.cell(row=row_excel,column=column_excel).value)

                    date_excel=excel_sheets.cell(row=1,column=column_excel).value
                    #根据时间（excel_date）和店铺名称（shop_id）来判断记录是否存在，如果存在更新原有记录，否则新建记录
                    if datadiff.objects.filter(date=date_excel, id_shop=shop_id).count():
                        sys_amount_base = datadiff.objects.get(date=date_excel, id_shop=shop_id).sys_amount
                        datadiff.objects.filter(date=date_excel, id_shop=shop_id).update(shop_amount=shopamount_excel,\
                                                                                         amount=sys_amount_base-shopamount_excel,\
                                                                                         true_amount=shopamount_excel)
                    else:
                        shop_diff=datadiff(shop_amount=shopamount_excel,date=date_excel,id_shop=shop_id,true_amount=shopamount_excel)
                        shop_diff.save()
                date_excel_str = date_excel.strftime("%Y-%m-%d")
        #数据导入后查询有差异的数据并将结果传递到前端渲染
        diff = datadiff.objects.exclude(amount=0) \
            .exclude(id_shop__sName__in=['元隆利嘉生活馆', '满洲里友谊商厦']) \
            .filter(id_shop__shopType__in=["D", "C"],
                    date__range=(date_excel_start, date_excel_str)) \
            .order_by("id_shop", "date")
        return render(request, 'checkdatanew.html', {"diff": diff, "shop_list": shop_list, "man_list": man_list})
        #批量导入人员信息
    elif excel_data_type=='saleflow':
        pass
    elif excel_data_type=="staff":
        for row_excel in range(2, excel_rows + 1):
                staff_name=excel_sheets.cell(row=row_excel,column=1).value
                staff_personal_phone=excel_sheets.cell(row=row_excel,column=2).value
                staff_c_phone=excel_sheets.cell(row=row_excel,column=3).value
                staff_email=excel_sheets.cell(row=row_excel,column=4).value
                staff_qq=excel_sheets.cell(row=row_excel,column=5).value
                staff_title=excel_sheets.cell(row=row_excel,column=6).value
                staff_address=excel_sheets.cell(row=row_excel,column=7).value
                if Managers.objects.filter(name=staff_name,personal_cellphone=staff_personal_phone).count() or \
                    Managers.objects.filter(name=staff_name,company_cellphone=staff_c_phone).count():
                    Managers.objects.filter(name=staff_name, personal_cellphone=staff_personal_phone).update(
                        name=staff_name,
                        personal_cellphone=staff_personal_phone,
                        company_cellphone=staff_c_phone,
                        qq=staff_qq,
                        email=staff_email,
                        title=staff_title,
                        address=staff_address
                    )
                else:
                    Managers.objects.create(name=staff_name,
                                            personal_cellphone=staff_personal_phone,
                                            company_cellphone=staff_c_phone,
                                            qq=staff_qq,
                                            email=staff_email,
                                            title=staff_title,
                                            address=staff_address
                                            )
                return render(request, 'shop.html', locals())

        #批量导入区域档案
    elif excel_data_type=='areainfo':
        for row_excel in range(2, excel_rows + 1):
                 #表格第二列为区域负责人姓名
                 manager_id=Managers.objects.get(name=(excel_sheets.cell(row=row_excel,column=2).value))
                 #创建区域信息
                 area_name=excel_sheets.cell(row=row_excel,column=1).value
                 Area.objects.update_or_create(name=area_name,manager=manager_id)
                 return HttpResponseRedirect("/")
        #批量导入门店档案

    elif excel_data_type=='shopinfo':
        for row_excel in range(2, excel_rows + 1):
                shop_sysname_excel=excel_sheets.cell(row=row_excel,column=1).value
                shop_reportname_excel=excel_sheets.cell(row=row_excel,column=2).value
                shop_manager_excel=Managers.objects.get(name=excel_sheets.cell(row=row_excel,column=3).value)
                shop_area_excel=Area.objects.get(name=excel_sheets.cell(row=row_excel,column=4).value)
                shop_type_excel=excel_sheets.cell(row=row_excel,column=5).value
                shop_state_excel=excel_sheets.cell(row=row_excel,column=6).value
                shop_address_excel=excel_sheets.cell(row=row_excel,column=7).value
                shop_mall_type_excel=excel_sheets.cell(row=row_excel,column=8).value
                shop_opendate_excel=excel_sheets.cell(row=row_excel,column=9).value
                shop_brand_excel=excel_sheets.cell(row=row_excel,column=12).value
                shop_contractBeginDate_excel=excel_sheets.cell(row=row_excel,column=10).value
                shop_contractEndDate_excel=excel_sheets.cell(row=row_excel,column=11).value
                ShopInfo.objects.update_or_create(
                    sysName=shop_sysname_excel,
                    sName=shop_reportname_excel,
                    managerId=shop_manager_excel,
                    areaId=shop_area_excel,
                    state=shop_state_excel,
                    mallType=shop_mall_type_excel,
                    shopType=shop_type_excel,
                    shopAddress=shop_address_excel,
                    contractBeginDate=shop_contractBeginDate_excel,
                    contractEndDate=shop_contractEndDate_excel,
                    openingDate=shop_opendate_excel,
                    shop_brand=shop_brand_excel
                )
                return render(request,'shop.html',locals())

def saleflowindb(request):
    try:
        saledatestart=request.POST['saledatestart']
        saledateEnd=request.POST['saledateend']
        excel_file = request.FILES['excelfile']
        saleflow.objects.filter(saleDate__range=[saledatestart, saledateEnd]).delete()
        filepath = os.path.abspath(os.path.dirname('__file__'))
        with open(filepath + '/static/upload/' + saledatestart + saledateEnd + '.xlsx', 'wb+') as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)
        destination.close()
        excel_data = openpyxl.load_workbook(filepath + '/static/upload/' + saledatestart +saledateEnd  + ".xlsx")

    except ValueError:
        pass


def fo(reqest):
    manager=Managers.objects.all()
    return render(reqest,'difftable.html',locals())


def read_diff_info(req):
    data=req.GET


    # 获取当前时间
    day_now = time.localtime()
    # 当月第一天日期
    day_begin = '%d-%02d-01' % (day_now.tm_year, day_now.tm_mon)
    wday, monthRange = calendar.monthrange(day_now.tm_year, day_now.tm_mon)
    # 当月最后一天日期
    day_end = '%d-%02d-%02d' % (day_now.tm_year, day_now.tm_mon, monthRange)
    limit=req.GET.get('limit')
    order=req.GET.get('order')
    offset=req.GET.get('offset')
    search=req.GET.get('search')

    if search:
        diff_info_all=datadiff.objects.filter(Q(id_shop__sysName__contains=search)|
                                              Q(id_shop__managerId__name__contains=search)|\
                                              Q(remark__contains=search)).filter(id_shop__shopType__in=['D','C']).\
                                       exclude(amount=0).exclude(id_shop__sName__in=['元隆利嘉生活馆','满洲里友谊商厦'])
    else:
        diff_info_all=datadiff.objects.exclude(amount=0)\
                           .exclude(id_shop__sName__in=['元隆利嘉生活馆','满洲里友谊商厦'])\
                           .filter(id_shop__shopType__in=["D", "C"],
                                   date__range=(day_begin,day_end))\
                           .order_by("id_shop", "-date")
    if not offset:
        offset=0
    if not limit:
        limit=15
    pageinator=Paginator(diff_info_all,limit)
    pages=int(int(offset) /int(limit)+1)
    data_json = {'total':diff_info_all.count(),'rows':[]}

    for diffinfo in pageinator.page(pages):

        data_json['rows'].append(
            {

             "sysName":diffinfo.id_shop.sysName,
             "manager":diffinfo.id_shop.managerId.name,
             "diffdate": diffinfo.date.strftime('%Y-%m-%d') if diffinfo.date else "",
             "shopamount":diffinfo.shop_amount,
             "sysamount":diffinfo.sys_amount,
             "amount":diffinfo.amount,
             "diff":diffinfo.diff,
             "remark":diffinfo.remark,
             "trueamount":diffinfo.true_amount,
            }
        )
    data_json_response = json.dumps(data_json)
    return HttpResponse(data_json_response)

